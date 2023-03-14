# @AUTHOR: DIOCAI
# DEVELOP TIME: 23/3/9 16:39
import openpyxl

wb = openpyxl.load_workbook('D:\\Personal\\example.xlsx')

# Getting sheets from the workbook

# print(wb.sheetnames)
# for sheet in wb:
#     print(sheet.title)
#
# mySheet = wb.create_sheet('mySheet')
# print(wb.sheetnames)

# sheet3 = wb.get_sheet_by_name('Sheet3')
# sheet4 = wb['mySheet']

# Getting cells from the sheets
ws = wb.active
# print(ws)
# print(ws['A1'])
# print(ws['A1'].value)

# c = ws['B1']
# print('Row {}, Column {} is {}'.format(c.row, c.column, c.value))
# print('Cell {} is {}\n'.format(c.coordinate, c.value))
#
# print(ws.cell(row=1, column=2))
# print(ws.cell(row=1, column=2).value)
#
# for i in range(1, 8, 2):
#     print(i, ws.cell(row=i, column=1).value)

# colC = ws['C']
# # print(colC)
# for c2 in colC:
#     print(c2.value)

row6 = ws[6]
col_range = ws['A:C']
row_range = ws[2:6]

# for col in col_range:
#     for cell in col:
#         print(cell.value)

# for row in row_range:
#     for cell in row:
#         print(cell.value)
