# @AUTHOR: DIOCAI
# DEVELOP TIME: 23/3/14 14:27

import os.path
from openpyxl import Workbook

rootdir = 'D:\\Personal\\学习\\Python\\newPythonProject\\udt'
# 遍历所有文件
for parent, dirnames, filenames in os.walk(rootdir):
   pass

udtList = {}

# print(dict.keys(tagDict))

for file in filenames:
    # udt格式
    udtDict = {'TYPE': '', 'AUTHOR': '', 'NAME': '', 'VERSION': '', 'STRUCT': {}}

    if 'txt' in file:
        print('read: ' + file)
        f = open(rootdir + '\\' + file, encoding='utf-8')

        lines = f.readlines()
        # udt 类型
        while True:
            temp_line = lines.pop(0)
            if 'TYPE' in temp_line:
                break
        udtDict['TYPE'] = temp_line[temp_line.find('"') + 1:temp_line.rfind('"')].strip()

        # udt 作者 名字 版本 非关键
        l = list(dict.keys(udtDict))
        for i in range(1, 4):
            while True:
                temp_line = lines.pop(0)
                if l[i] in temp_line:
                    break
            udtDict[l[i]] = temp_line[temp_line.find(':') + 1:temp_line.rfind('\n')].strip()

        # print(udtDict['TYPE'])
        # print(udtDict['AUTHOR'])
        # print(udtDict['NAME'])
        # print(udtDict['VERSION'])

        # 读变量
        while len(lines) > 0:
            temp_line = lines.pop(0)
            if 'END_TYPE' in temp_line:
                print('finished: ' + file)
                break
            elif ':' not in temp_line:
                continue
            # 变量格式
            tagDict = {'OFFSET': 0, 'NAME': '', 'TYPE': '', 'REMARK': ''}
            start_name = 0
            end_name = temp_line.find(':')
            start_type = end_name + 1
            end_type = temp_line.find(';')
            tagDict['NAME'] = temp_line[start_name:end_name].strip()
            tagDict['TYPE'] = temp_line[start_type:end_type].strip()
            if '//' in temp_line:
                tagDict['REMARK'] = temp_line[temp_line.find('//') + 2: temp_line.find('\n')].strip()
            udtDict['STRUCT'][tagDict['NAME']] = tagDict

        f.close()
    if udtDict['TYPE'] != '':
        udtList[udtDict['TYPE']] = udtDict

print(udtList)
print('udt数量: ' + str(len(udtList)))

'''
以下是生成excel的部分
'''
print('\n开始生成excel\n')
wb = Workbook()
for udt in udtList:
    print('生成Sheet: ' + udt)
    ws = wb.create_sheet(udt, 0)
    ws.append(list(tagDict))
    for tag in udtList[udt]['STRUCT']:
        l = []
        temp = udtList[udt]['STRUCT'][tag]
        for fd in tagDict:
            l.append(temp[fd])
        ws.append(l)
    # r = 1
    # for row in udt:
    #     if (row != 'STRUCT'):
    #         ws.cell(row=r, column=1).value = row
    #         ws.cell(row=r, column=2).value = udt[row]
    #     else:
    #         for filed in udt[row]:
    #             c = 1
    #             for i in udt[row][filed]:
    #                 ws.cell(row=r, column=c).value = udt[row][filed][i]
    #     r += 1
        # ws.append(list(udt[row]))

wb.save(rootdir + '\\' + 'test.xlsx')
